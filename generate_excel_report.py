"""
Generate Excel report from JSON analysis results
Creates simplified multi-sheet workbook for validation
"""
import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_excel_report(json_path: Path):
    """Generate Excel report from JSON results"""

    # Load JSON data
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    stats = data['statistics']
    results = data['results']

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Executive Summary
    ws_summary = wb.create_sheet("Executive Summary")
    create_summary_sheet(ws_summary, data, stats, results)

    # Sheet 2: Detected Posts
    ws_detected = wb.create_sheet("Nicotine Detections")
    create_detections_sheet(ws_detected, results)

    # Sheet 3: All Posts Database
    ws_all = wb.create_sheet("All Posts")
    create_all_posts_sheet(ws_all, results)

    # Save workbook
    output_path = json_path.parent / f"report_{json_path.stem}.xlsx"
    wb.save(output_path)

    print(f"\n[SUCCESS] Excel report generated: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  - Executive Summary")
    print(f"  - Nicotine Detections ({stats['nicotine_detected']} posts)")
    print(f"  - All Posts ({stats['total_posts']} posts)")

    return output_path


def create_summary_sheet(ws, data, stats, results):
    """Create executive summary sheet"""

    # Title
    ws['A1'] = "IQOS GEORGIA SOCIAL INTELLIGENCE ANALYSIS"
    ws['A1'].font = Font(size=16, bold=True, color="00B8A9")
    ws['A2'] = f"Generated: {data['metadata']['generated_at']}"
    ws['A2'].font = Font(size=10, italic=True)

    # Key Metrics
    row = 4
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    metrics = [
        ("Total Posts Analyzed", stats['total_posts']),
        ("Successful Analysis", stats['successful']),
        ("Failed", stats['failed']),
        ("", ""),
        ("Nicotine Detected", stats['nicotine_detected']),
        ("Detection Rate", f"{stats['nicotine_detected']/stats['total_posts']*100:.1f}%"),
    ]

    for label, value in metrics:
        if label:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        else:
            row += 1

    # Category Breakdown
    row += 1
    ws[f'A{row}'] = "DETECTION BY CATEGORY"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    for category, count in stats.get('by_category', {}).items():
        ws[f'A{row}'] = category
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{count/stats['nicotine_detected']*100:.1f}%" if stats['nicotine_detected'] > 0 else "0%"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Influencer Overview
    row += 2
    ws[f'A{row}'] = "TOP INFLUENCERS WITH DETECTIONS"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    ws[f'A{row}'] = "Username"
    ws[f'B{row}'] = "Detections"
    ws[f'C{row}'] = "Total Posts"
    ws[f'D{row}'] = "Detection Rate"

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row += 1

    # Calculate per-influencer stats
    influencer_stats = {}
    for r in results:
        username = r['username']
        if username not in influencer_stats:
            influencer_stats[username] = {'total': 0, 'detected': 0}
        influencer_stats[username]['total'] += 1
        if r['analysis']['nicotine_detection']['detected']:
            influencer_stats[username]['detected'] += 1

    # Sort by detection count
    sorted_influencers = sorted(influencer_stats.items(),
                               key=lambda x: x[1]['detected'],
                               reverse=True)

    for username, counts in sorted_influencers[:10]:  # Top 10
        if counts['detected'] > 0:
            ws[f'A{row}'] = f"@{username}"
            ws[f'B{row}'] = counts['detected']
            ws[f'C{row}'] = counts['total']
            ws[f'D{row}'] = f"{counts['detected']/counts['total']*100:.1f}%"
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_detections_sheet(ws, results):
    """Create sheet with detected posts"""

    # Headers
    headers = [
        "Username", "Post URL", "Date", "Type",
        "Detected", "Confidence", "Category", "Specific Brand",
        "Visual Evidence", "Caption Evidence", "Comments Evidence",
        "Usage Context", "Usage Type"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="00B8A9", end_color="00B8A9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Data rows - only detected posts
    row = 2
    for r in results:
        nd = r['analysis']['nicotine_detection']

        if not nd['detected']:
            continue

        # Extract product info
        products = nd.get('products', [])
        category = products[0]['category'] if products else ""
        brand = products[0].get('specific_brand', '') if products else ""

        # Extract evidence - handle both list of strings and single strings
        visual_list = nd['detection_evidence'].get('visual', [])
        caption_list = nd['detection_evidence'].get('caption', [])
        comments_list = nd['detection_evidence'].get('comments', [])

        # Ensure we're joining strings, not characters
        visual_ev = " | ".join(str(v) for v in visual_list if v and v != "not_mentioned")[:300]
        caption_ev = " | ".join(str(c) for c in caption_list if c and c != "not_mentioned")[:300]
        comments_ev = " | ".join(str(c) for c in comments_list if c and c != "not_mentioned")[:300]

        # Write row
        ws.cell(row=row, column=1).value = f"@{r['username']}"
        ws.cell(row=row, column=2).value = r['url']
        ws.cell(row=row, column=2).hyperlink = r['url']
        ws.cell(row=row, column=2).font = Font(color="0563C1", underline="single")
        ws.cell(row=row, column=3).value = r['timestamp'][:10]
        ws.cell(row=row, column=4).value = r['post_type']
        ws.cell(row=row, column=5).value = "YES"
        ws.cell(row=row, column=6).value = nd['confidence']
        ws.cell(row=row, column=7).value = category
        ws.cell(row=row, column=8).value = brand
        ws.cell(row=row, column=9).value = visual_ev
        ws.cell(row=row, column=10).value = caption_ev
        ws.cell(row=row, column=11).value = comments_ev
        ws.cell(row=row, column=12).value = nd.get('usage_context', '')
        ws.cell(row=row, column=13).value = nd.get('usage_type', '')

        # Wrap text for evidence columns
        for col in [9, 10, 11]:
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        row += 1

    # Column widths
    widths = [20, 15, 12, 10, 10, 12, 15, 20, 50, 50, 50, 20, 20]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def create_all_posts_sheet(ws, results):
    """Create sheet with all posts"""

    # Headers
    headers = [
        "Username", "Post URL", "Date", "Type", "Likes", "Comments",
        "Nicotine Detected", "Confidence", "Category",
        "Caption", "Hashtags", "Primary Language"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    # Data rows - all posts
    row = 2
    for r in results:
        nd = r['analysis']['nicotine_detection']
        meta = r['analysis']['metadata']

        # Get post details from original data (if available)
        # For now, use what's in the analysis

        ws.cell(row=row, column=1).value = f"@{r['username']}"
        ws.cell(row=row, column=2).value = r['url']
        ws.cell(row=row, column=2).hyperlink = r['url']
        ws.cell(row=row, column=2).font = Font(color="0563C1", underline="single")
        ws.cell(row=row, column=3).value = r['timestamp'][:10]
        ws.cell(row=row, column=4).value = r['post_type']
        ws.cell(row=row, column=5).value = "N/A"  # Would need from original post data
        ws.cell(row=row, column=6).value = meta.get('comment_count_analyzed', 0)
        ws.cell(row=row, column=7).value = "YES" if nd['detected'] else "NO"
        ws.cell(row=row, column=8).value = nd['confidence'] if nd['detected'] else ""

        # Category
        products = nd.get('products', [])
        category = products[0]['category'] if products else ""
        ws.cell(row=row, column=9).value = category

        # Color code detection column
        if nd['detected']:
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.cell(row=row, column=10).value = ""  # Would need caption from original
        ws.cell(row=row, column=11).value = ""  # Would need hashtags from original
        ws.cell(row=row, column=12).value = meta.get('primary_language', '')

        row += 1

    # Column widths
    widths = [20, 15, 12, 10, 10, 10, 15, 12, 15, 50, 30, 15]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


if __name__ == '__main__':
    # Find most recent JSON file
    reports_dir = Path('output/reports')
    json_files = sorted(reports_dir.glob('analysis_results_*.json'),
                       key=lambda p: p.stat().st_mtime,
                       reverse=True)

    if json_files:
        print(f"\n[GENERATING EXCEL REPORT]")
        print(f"Source: {json_files[0].name}")
        generate_excel_report(json_files[0])
    else:
        print("No JSON result files found in output/reports/")
